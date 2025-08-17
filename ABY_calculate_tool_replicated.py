from openpyxl import load_workbook
from Bio import SeqIO

def count_overlapping(seq, sub):
    """Count overlapping occurrences of sub in seq."""
    count = start = 0
    while True:
        start = seq.find(sub, start)
        if start == -1:
            return count
        count += 1
        start += 1  # move by one to allow overlaps

def main():
    species_name = input("Enter the species name: ").strip()
    protein_for_species_file = input("Enter the path to your protein list for certain species file: ").strip()
    peptide_list_file = input("Enter the path to peptides .xlsx file: ").strip()
    
    # Load peptide list
    wb = load_workbook(peptide_list_file)
    ws = wb.active
    
    peptides = []
    ec50_dict = {}
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        seq = str(row[6]).strip().upper() if row[6] else None  # Column G = index 6
        ec50 = row[4]  # Column E = index 4
        if seq:
            peptides.append(seq)
            ec50_dict[seq] = ec50
    
    # Load protein sequences
    protein_records = list(SeqIO.parse(protein_for_species_file, 'fasta'))
    
    output_file = f"{species_name.replace(' ', '_')}_results.txt"
    
    with open(output_file, 'w', encoding='utf-8') as outfile:
        
        for protein in protein_records:
            protein_seq = str(protein.seq).upper()
            N = len(protein_seq)
            
            a_total = 0
            sum_ai_over_ec50 = 0
            matched_peptides = []
            
            for pep in peptides:
                count = count_overlapping(protein_seq, pep)
                if count > 0:
                    matched_peptides.append(pep)
                    a_total += count
                    
                    ec50_value = ec50_dict.get(pep)
                    if ec50_value and ec50_value != 0:
                        sum_ai_over_ec50 += count / ec50_value
            
            k = len(set(matched_peptides))
            A = a_total / N if N > 0 else 0
            B = sum_ai_over_ec50 / N if N > 0 else 0
            
            outfile.write(f"Species: {species_name}\n")
            outfile.write(f"Protein ID: {protein.id}\n")
            outfile.write(f"Description: {protein.description}\n")
            outfile.write(f"Length (N): {N}\n")
            outfile.write(f"a (total fragments): {a_total}\n")
            outfile.write(f"k (unique fragments): {k}\n")
            outfile.write(f"A: {round(A, 6)}\n")
            outfile.write(f"B (µM^-1): {round(B, 6)}\n")
            outfile.write(f"Matched peptides: {';'.join(set(matched_peptides))}\n")
            outfile.write("\n" + "-"*40 + "\n\n")
    
    print(f"Results saved to {output_file}")

if __name__ == '__main__':
    main()
